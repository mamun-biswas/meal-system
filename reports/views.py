from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum
from django.conf import settings
from django.views.decorators.clickjacking import xframe_options_sameorigin
from decimal import Decimal
from accounts.models import Member
from accounts.views import log_action
from accounts.month_helpers import get_active_my, months_list, years_list
from finance.models import Deposit, Expense
from meals.models import MealMark, MonthlySettings
from meals.views import compute_month_stats
import datetime, calendar, json, io


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _bn_num(v):
    """Convert ASCII digits in a string/number to Bangla digits."""
    return str(v).replace('0','০').replace('1','১').replace('2','২') \
                 .replace('3','৩').replace('4','৪').replace('5','৫') \
                 .replace('6','৬').replace('7','৭').replace('8','৮').replace('9','৯')

BN_MONTHS = ['জানুয়ারি','ফেব্রুয়ারি','মার্চ','এপ্রিল','মে','জুন',
             'জুলাই','আগস্ট','সেপ্টেম্বর','অক্টোবর','নভেম্বর','ডিসেম্বর']

def _bn_month_year(month, year):
    return f"{BN_MONTHS[month-1]} {_bn_num(year)}"

def _member_display_name(m, lang='en'):
    if lang == 'bn':
        return m.name_bn if m.name_bn else m.name
    return m.name


# ---------------------------------------------------------------------------
# Existing views
# ---------------------------------------------------------------------------

@login_required
def report_monthly(request):
    month, year = get_active_my(request)
    stats  = compute_month_stats(month, year)
    today  = datetime.date.today()
    # Preloaded once for the whole 6-month trend loop below, instead of
    # a fresh DayConfig + MealCountSettings query per MealMark row.
    day_config_map, weights = MealMark.preload_calc_context(request.user.member.mess)
    trend  = []
    for i in range(5, -1, -1):
        mn = ((today.month - 1 - i) % 12) + 1
        yr = today.year if i <= today.month - 1 else today.year - 1
        te  = MealMark.objects.filter(date__month=mn, date__year=yr).select_related('member__mess')
        eff = sum((mk.effective_count(day_config_map, weights) for mk in te), Decimal('0'))
        exp = Expense.objects.filter(date__month=mn, date__year=yr
              ).aggregate(t=Sum('amount'))['t'] or 0
        dep = Deposit.objects.filter(date__month=mn, date__year=yr
              ).aggregate(t=Sum('amount'))['t'] or 0
        trend.append({
            'label': f"{calendar.month_abbr[mn]}'{str(yr)[-2:]}",
            'eff': float(eff), 'exp': float(exp), 'dep': float(dep),
        })
    return render(request, 'reports/monthly.html', {
        'stats': stats, 'months': months_list(), 'years': years_list(year),
        'month': month, 'year': year,
        'month_name': calendar.month_name[month],
        'trend_json': json.dumps(trend),
    })


@login_required
def report_daily(request):
    today    = datetime.date.today()
    date_str = request.GET.get('date', str(today))
    try:
        date = datetime.date.fromisoformat(date_str)
    except Exception:
        date = today
    marks  = list(MealMark.objects.filter(date=date).select_related('member__mess', 'marked_by'))
    exps   = Expense.objects.filter(date=date).select_related('bought_by', 'added_by')
    deps   = Deposit.objects.filter(date=date).select_related('member', 'added_by')
    day_config_map, weights = MealMark.preload_calc_context(request.user.member.mess, dates=[date])
    total_meals = sum((mk.effective_count(day_config_map, weights) for mk in marks), Decimal('0'))
    total_exp   = exps.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    total_dep   = deps.aggregate(t=Sum('amount'))['t'] or Decimal('0')
    return render(request, 'reports/daily.html', {
        'date': date, 'marks': marks, 'expenses': exps, 'deposits': deps,
        'total_meals': total_meals,
        'total_expense': total_exp,
        'total_deposit': total_dep,
    })


@login_required
def export_csv(request):
    member = request.user.member
    if not member.has_perm_code('export_data'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    month, year  = get_active_my(request)
    report_type  = request.GET.get('type', 'monthly')
    import csv
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = (
        f'attachment; filename="meal_{report_type}_{month}_{year}.csv"'
    )
    writer = csv.writer(response)
    if report_type == 'monthly':
        stats = compute_month_stats(month, year)
        writer.writerow(['Member', 'Room', 'Raw Meals', 'Eff. Meals',
                         'Meal Cost', 'Cook Cost', 'Total Cost', 'Deposit', 'Balance'])
        for s in stats['per_member']:
            writer.writerow([
                s['member'].name, s['member'].room_number,
                s['raw_meals'], s['eff_meals'],
                round(float(s['cost']) - float(s['cook_cost']), 2),
                s['cook_cost'], s['cost'], s['deposit'], s['balance'],
            ])
        writer.writerow([])
        writer.writerow(['Meal Rate', stats['meal_rate']])
        writer.writerow(['Total Expense', stats['total_exp']])
        writer.writerow(['Fund Balance', stats['fund_balance']])
    elif report_type == 'deposits':
        deps = Deposit.objects.filter(date__month=month, date__year=year
               ).select_related('member', 'added_by').order_by('date')
        writer.writerow(['Date', 'Member', 'Amount', 'Method', 'Note', 'Added By'])
        for d in deps:
            writer.writerow([d.date, d.member.name, d.amount, d.method,
                             d.note, d.added_by.name if d.added_by else ''])
    elif report_type == 'expenses':
        exps = Expense.objects.filter(date__month=month, date__year=year
               ).select_related('bought_by', 'added_by').order_by('date')
        writer.writerow(['Date', 'Amount', 'Category', 'Description', 'Bought By', 'Added By'])
        for e in exps:
            writer.writerow([e.date, e.amount, e.get_category_display(),
                             e.description,
                             e.bought_by.name if e.bought_by else '',
                             e.added_by.name if e.added_by else ''])
    log_action(request.user.member,
               f'Exported CSV ({report_type}): {calendar.month_name[month]} {year}',
               request=request)
    return response


# ---------------------------------------------------------------------------
# Export: Excel (English)
# ---------------------------------------------------------------------------

@login_required
def export_excel_en(request):
    member = request.user.member
    if not member.has_perm_code('export_data'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    month = int(request.GET.get('month', get_active_my(request)[0]))
    year  = int(request.GET.get('year',  get_active_my(request)[1]))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter

    stats = compute_month_stats(month, year)
    mess_name = member.mess.name
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = f"{calendar.month_name[month]} {year}"

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title rows ──────────────────────────────────────────────────────────
    ws.merge_cells('A1:J1')
    t1 = ws['A1']
    t1.value = f'{mess_name} — Monthly Financial Report'
    t1.font  = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    t1.fill  = PatternFill('solid', fgColor='1E2235')
    t1.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:J2')
    t2 = ws['A2']
    t2.value = f"{settings.APP_NAME} · Mess ID: {member.mess.code} · {calendar.month_name[month]} {year}"
    t2.font  = Font(name='Calibri', bold=True, size=11, color='AAAACC')
    t2.fill  = PatternFill('solid', fgColor='1E2235')
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Summary row ─────────────────────────────────────────────────────────
    ws.merge_cells('A3:B3')
    ws['A3'].value = f"Meal Rate: ৳{stats['meal_rate']}"
    ws.merge_cells('C3:D3')
    ws['C3'].value = f"Total Expense: ৳{stats['total_exp']}"
    ws.merge_cells('E3:F3')
    ws['E3'].value = f"Total Deposit: ৳{stats['total_dep']}"
    ws.merge_cells('G3:H3')
    ws['G3'].value = f"Fund Balance: ৳{stats['fund_balance']}"
    ws.merge_cells('I3:J3')
    ws['I3'].value = f"Cook Cost/Member: ৳{stats['cook_cost']}"
    for col in range(1, 11):
        c = ws.cell(3, col)
        c.font = Font(name='Calibri', bold=True, size=9, color='333333')
        c.fill = PatternFill('solid', fgColor='EEF0FF')
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 16

    # ── Headers ─────────────────────────────────────────────────────────────
    headers = ['#', 'Member Name', 'Room', 'Raw Meals', 'Eff. Meals',
               'Meal Cost (৳)', 'Cook Cost (৳)', 'Total Cost (৳)', 'Deposit (৳)', 'Balance (৳)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font      = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill      = PatternFill('solid', fgColor='4F46E5')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = border
    ws.row_dimensions[4].height = 22

    # ── Data rows ───────────────────────────────────────────────────────────
    for i, s in enumerate(stats['per_member'], 1):
        row = i + 4
        bal = float(s['balance'])
        row_data = [
            i,
            s['member'].name,
            s['member'].room_number or '—',
            float(s['raw_meals']),
            float(s['eff_meals']),
            round(float(s['cost']) - float(s['cook_cost']), 2),
            float(s['cook_cost']),
            float(s['cost']),
            float(s['deposit']),
            round(bal, 2),
        ]
        fill_color = 'FFFFFF' if i % 2 == 0 else 'F8F9FF'
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row, col, val)
            c.font   = Font(name='Calibri', size=10)
            c.fill   = PatternFill('solid', fgColor=fill_color)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col == 2:
                c.alignment = Alignment(horizontal='left', vertical='center')
            if col == 10:
                c.font = Font(name='Calibri', size=10, bold=True,
                              color='1A7A1A' if bal >= 0 else 'CC0000')
        ws.row_dimensions[row].height = 18

    # ── Totals footer ───────────────────────────────────────────────────────
    foot_row = len(stats['per_member']) + 5
    ws.merge_cells(f'A{foot_row}:C{foot_row}')
    ws.cell(foot_row, 1, 'TOTAL')
    for col in range(1, 11):
        c = ws.cell(foot_row, col)
        c.font   = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
        c.fill   = PatternFill('solid', fgColor='1E2235')
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(foot_row, 5, float(stats['total_eff']))
    ws.cell(foot_row, 8, float(stats['total_exp']))
    ws.cell(foot_row, 9, float(stats['total_dep']))
    ws.row_dimensions[foot_row].height = 20

    # ── Print date ──────────────────────────────────────────────────────────
    note_row = foot_row + 1
    ws.merge_cells(f'A{note_row}:J{note_row}')
    ws.cell(note_row, 1, f"Generated: {datetime.date.today().strftime('%d %B %Y')}")
    ws.cell(note_row, 1).font = Font(name='Calibri', size=9, color='888888', italic=True)
    ws.cell(note_row, 1).alignment = Alignment(horizontal='right')

    # ── Column widths ───────────────────────────────────────────────────────
    col_widths = [5, 22, 8, 12, 12, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Print setup ─────────────────────────────────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"meal_report_en_{month}_{year}.xlsx"
    log_action(request.user.member,
               f'Exported Excel (English): {calendar.month_name[month]} {year}',
               request=request)
    response = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# Export: Excel (Bangla)
# ---------------------------------------------------------------------------

@login_required
def export_excel_bn(request):
    member = request.user.member
    if not member.has_perm_code('export_data'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    month = int(request.GET.get('month', get_active_my(request)[0]))
    year  = int(request.GET.get('year',  get_active_my(request)[1]))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    stats = compute_month_stats(month, year)
    mess_name = member.mess.name
    wb    = openpyxl.Workbook()
    ws    = wb.active
    ws.title = _bn_month_year(month, year)

    thin   = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    BN_FONT = 'Hind Siliguri'  # fallback: Arial Unicode MS

    def bf(size=10, color='000000'):
        return Font(name=BN_FONT, bold=True, size=size, color=color)

    # ── Title row (now includes mess name) ───────────────────────────────────
    ws.merge_cells('A1:I1')
    t1 = ws['A1']
    t1.value = f'{mess_name} — মাসিক হিসাব রিপোর্ট'
    t1.font  = bf(14, 'FFFFFF')
    t1.fill  = PatternFill('solid', fgColor='1E2235')
    t1.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 26

    ws.merge_cells('A2:I2')
    t2 = ws['A2']
    t2.value = _bn_month_year(month, year)
    t2.font  = bf(11, 'AAAACC')
    t2.fill  = PatternFill('solid', fgColor='1E2235')
    t2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # ── Summary row ─────────────────────────────────────────────────────────
    summaries = [
        ('মিল রেট', f"৳{_bn_num(stats['meal_rate'])}"),
        ('মোট বাজার', f"৳{_bn_num(stats['total_exp'])}"),
        ('মোট জমা', f"৳{_bn_num(stats['total_dep'])}"),
        ('ফান্ড ব্যালেন্স', f"৳{_bn_num(stats['fund_balance'])}"),
        ('রান্না খরচ/জন', f"৳{_bn_num(stats['cook_cost'])}"),
    ]
    col = 1
    for label, val in summaries:
        ws.cell(3, col, f"{label}: {val}")
        c = ws.cell(3, col)
        c.font = bf(9, '333333')
        c.fill = PatternFill('solid', fgColor='EEF0FF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
        col += 2
    ws.row_dimensions[3].height = 16

    # ── Headers ─────────────────────────────────────────────────────────────
    headers = ['ক্র:', 'নাম', 'মোট মিল', 'মিল খরচ (৳)', 'রান্না খরচ (৳)',
               'মোট খরচ (৳)', 'জমা (৳)', 'পাবে (৳)', 'দিবে (৳)']
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font      = bf(10, 'FFFFFF')
        c.fill      = PatternFill('solid', fgColor='4F46E5')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = border
    ws.row_dimensions[4].height = 22

    # ── Data rows ───────────────────────────────────────────────────────────
    for i, s in enumerate(stats['per_member'], 1):
        row = i + 4
        bal  = float(s['balance'])
        pabe = _bn_num(f"{abs(bal):.2f}") if bal >= 0 else '—'
        dibe = _bn_num(f"{abs(bal):.2f}") if bal < 0  else '—'
        row_data = [
            _bn_num(i),
            _member_display_name(s['member'], 'bn'),
            _bn_num(f"{float(s['eff_meals']):.1f}"),
            _bn_num(f"{round(float(s['cost']) - float(s['cook_cost']), 2):.2f}"),
            _bn_num(f"{float(s['cook_cost']):.0f}"),
            _bn_num(f"{float(s['cost']):.2f}"),
            _bn_num(f"{float(s['deposit']):.0f}"),
            pabe,
            dibe,
        ]
        fill_color = 'FFFFFF' if i % 2 == 0 else 'F8F9FF'
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row, col, val)
            c.font   = bf(10, '000000')
            c.fill   = PatternFill('solid', fgColor=fill_color)
            c.border = border
            c.alignment = Alignment(horizontal='center', vertical='center')
            if col == 2:
                c.alignment = Alignment(horizontal='left', vertical='center')
            if col == 8:
                c.font = bf(10, '1A7A1A')
            if col == 9:
                c.font = bf(10, 'CC0000')
        ws.row_dimensions[row].height = 18

    # ── Totals footer ───────────────────────────────────────────────────────
    foot_row = len(stats['per_member']) + 5
    ws.merge_cells(f'A{foot_row}:B{foot_row}')
    ws.cell(foot_row, 1, 'মোট')
    for col in range(1, 10):
        c = ws.cell(foot_row, col)
        c.font   = bf(10, 'FFFFFF')
        c.fill   = PatternFill('solid', fgColor='1E2235')
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(foot_row, 3, _bn_num(f"{float(stats['total_eff']):.1f}"))
    ws.cell(foot_row, 6, _bn_num(f"{float(stats['total_exp']):.0f}"))
    ws.cell(foot_row, 7, _bn_num(f"{float(stats['total_dep']):.0f}"))
    ws.row_dimensions[foot_row].height = 20

    # ── Print date ──────────────────────────────────────────────────────────
    note_row = foot_row + 1
    today = datetime.date.today()
    print_date = f"{_bn_num(today.day)} {BN_MONTHS[today.month-1]} {_bn_num(today.year)}"
    ws.merge_cells(f'A{note_row}:I{note_row}')
    ws.cell(note_row, 1, f"মুদ্রণের তারিখ: {print_date}")
    ws.cell(note_row, 1).font = bf(9, '888888')
    ws.cell(note_row, 1).alignment = Alignment(horizontal='right')

    # ── Column widths ───────────────────────────────────────────────────────
    col_widths = [6, 22, 12, 16, 16, 16, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Print setup ─────────────────────────────────────────────────────────
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize   = ws.PAPERSIZE_A4
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"meal_report_bn_{month}_{year}.xlsx"
    log_action(request.user.member,
               f'Exported Excel (Bangla): {calendar.month_name[month]} {year}',
               request=request)
    response = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# Export: PDF (English)
# ---------------------------------------------------------------------------

@login_required
def export_pdf_en_preview(request):
    """HTML rendering of the English monthly report, styled to look like
    the PDF, shown inside the preview modal's iframe via srcdoc.

    Deliberately NOT a PDF: some browsers are configured (Chrome's
    "Download PDFs instead of automatically opening them") to force a
    download for ANY application/pdf response — including one loaded
    into an iframe via a blob: URL, or opened directly via window.open()
    despite an `inline` Content-Disposition — which makes viewing an
    actual PDF unreliable both for the preview modal AND for the Print
    button. Bangla's report sidesteps this entirely by being plain HTML
    with the browser's own print dialog triggered via a script; this view
    does the same for English, used for both the preview iframe
    (?preview=1, no script) and the Print button (?print=1, with the
    auto-print script). The real PDF (via export_pdf_en) is only ever
    produced when the user clicks Download.
    """
    member = request.user.member
    if not member.has_perm_code('export_data'):
        return HttpResponse('Access denied.', status=403)
    month = int(request.GET.get('month', get_active_my(request)[0]))
    year  = int(request.GET.get('year',  get_active_my(request)[1]))
    is_print = request.GET.get('print') == '1'
    auto_print_script = """<script>
window.onload = function() {
  setTimeout(function() { window.print(); }, 700);
};
</script>""" if is_print else ''
    stats = compute_month_stats(month, year)
    mess_name = member.mess.name
    now = datetime.datetime.now()

    fund_bal = float(stats['fund_balance'])
    fund_color = '#0F8A3E' if fund_bal >= 0 else '#D42A2A'

    rows_html = ''
    total_will_pay = Decimal('0')
    total_will_receive = Decimal('0')
    for i, s in enumerate(stats['per_member'], 1):
        bal = float(s['balance'])
        if bal < 0:
            will_pay_html = f'<td class="pay">Tk {abs(bal):.2f}</td>'
            will_receive_html = '<td class="ctr">—</td>'
            total_will_pay += abs(s['balance'])
        else:
            will_pay_html = '<td class="ctr">—</td>'
            will_receive_html = f'<td class="rcv">Tk {bal:.2f}</td>'
            total_will_receive += s['balance']
        bg = '#F7F8FF' if i % 2 == 0 else '#fff'
        rows_html += (
            f'<tr style="background:{bg}">'
            f'<td class="ctr">{i}</td>'
            f'<td class="left">{s["member"].name}</td>'
            f'<td class="ctr">{s["member"].room_number or "—"}</td>'
            f'<td class="ctr">{float(s["eff_meals"]):.1f}</td>'
            f'<td class="ctr">Tk {round(float(s["cost"])-float(s["cook_cost"]),2):.2f}</td>'
            f'<td class="ctr">Tk {float(s["cook_cost"]):.0f}</td>'
            f'<td class="ctr">Tk {float(s["cost"]):.2f}</td>'
            f'<td class="ctr">Tk {float(s["deposit"]):.0f}</td>'
            f'{will_pay_html}{will_receive_html}'
            f'</tr>'
        )

    html = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Monthly Report — {calendar.month_name[month]} {year}</title>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:Arial,Helvetica,sans-serif;font-size:10pt;color:#161A2E;background:#fff;padding:10mm 12mm}}
.hdr{{background:#161A2E;color:#fff;padding:10px 14px 8px;border-radius:4px;display:flex;justify-content:space-between;align-items:flex-end}}
.hdr .brand{{font-size:17pt;font-weight:bold}}
.hdr .sub{{font-size:8.5pt;color:#B9BAE0;margin-top:2px}}
.hdr .title{{font-size:12.5pt;font-weight:bold;text-align:right}}
.hdr .period{{font-size:9.5pt;color:#B9BAE0;text-align:right;margin-top:2px}}
.accent{{height:2.4px;background:#4F46E5}}
.meta{{display:flex;justify-content:space-between;font-size:7.8pt;color:#8A8AA0;margin:6px 0 10px;padding-bottom:6px;border-bottom:1px solid #E1E2F0}}
.section{{font-size:10.5pt;font-weight:bold;color:#161A2E;margin:8px 0 6px}}
.cards{{display:flex;gap:0;border:0.6px solid #E1E2F0;background:#F0F1FE;border-radius:2px;margin-bottom:12px}}
.card{{flex:1;text-align:center;padding:8px 4px;border-right:0.6px solid #E1E2F0;border-top:2.2px solid var(--c)}}
.card:last-child{{border-right:none}}
.card .l{{font-size:7.6pt;color:#8A8AA0;text-transform:uppercase}}
.card .v{{font-size:13pt;font-weight:bold;margin-top:3px}}
table{{width:100%;border-collapse:collapse;font-size:8.8pt}}
th{{background:#4F46E5;color:#fff;padding:6px 4px;text-align:center;font-size:8.6pt}}
td{{padding:5px 4px;border:0.35px solid #E1E2F0;text-align:center}}
td.left{{text-align:left;padding-left:6px}}
td.ctr{{text-align:center}}
td.pay{{color:#D42A2A;font-weight:bold}}
td.rcv{{color:#0F8A3E;font-weight:bold}}
tfoot td{{background:#161A2E;color:#fff;font-weight:bold;padding:6px 4px}}
.note{{font-size:7.6pt;color:#8A8AA0;font-style:italic;margin-top:8px}}
.sigs{{display:flex;justify-content:space-around;margin-top:22px}}
.sig{{text-align:center;font-size:8pt;color:#161A2E;border-top:1px solid #161A2E;padding-top:4px;min-width:120px}}
.foot{{margin-top:14px;padding-top:6px;border-top:0.5px solid #E1E2F0;display:flex;justify-content:space-between;font-size:7pt;color:#8A8AA0}}
</style>
</head>
<body>
  <div class="hdr">
    <div><div class="brand">{mess_name}</div><div class="sub">{settings.APP_NAME} · Mess ID: {member.mess.code}</div></div>
    <div><div class="title">MONTHLY FINANCIAL REPORT</div><div class="period">{calendar.month_name[month]} {year}</div></div>
  </div>
  <div class="accent"></div>
  <div class="meta">
    <span>Prepared by: <b>{member.name}</b> ({member.get_role_display()})</span>
    <span>Members: <b>{len(stats['per_member'])}</b></span>
    <span>Generated: {now.strftime('%d %b %Y, %I:%M %p')}</span>
  </div>

  <div class="section">FINANCIAL SUMMARY</div>
  <div class="cards">
    <div class="card" style="--c:#4F46E5"><div class="l">Meal Rate</div><div class="v">Tk {stats['meal_rate']}</div></div>
    <div class="card" style="--c:#D42A2A"><div class="l">Total Expense</div><div class="v">Tk {stats['total_exp']:.0f}</div></div>
    <div class="card" style="--c:#0F8A3E"><div class="l">Total Deposit</div><div class="v">Tk {stats['total_dep']:.0f}</div></div>
    <div class="card" style="--c:#7C3AED"><div class="l">Effective Meals</div><div class="v">{stats['total_eff']:.1f}</div></div>
    <div class="card" style="--c:#B45309"><div class="l">Cook Cost / Member</div><div class="v">Tk {stats['cook_cost']}</div></div>
    <div class="card" style="--c:{fund_color}"><div class="l">Fund Balance</div><div class="v" style="color:{fund_color}">Tk {fund_bal:.0f}</div></div>
  </div>

  <div class="section">MEMBER-WISE BREAKDOWN</div>
  <table>
    <thead><tr>
      <th>#</th><th>Member Name</th><th>Room</th><th>Total Meal</th>
      <th>Meal Cost (Tk)</th><th>Cook Cost (Tk)</th><th>Total Cost (Tk)</th>
      <th>Deposit (Tk)</th><th>Will Pay (Tk)</th><th>Will Receive (Tk)</th>
    </tr></thead>
    <tbody>{rows_html}</tbody>
    <tfoot><tr>
      <td>TOTAL</td><td></td><td></td>
      <td>{float(stats['total_eff']):.1f}</td><td></td><td></td>
      <td>Tk {float(stats['total_exp']):.0f}</td>
      <td>Tk {float(stats['total_dep']):.0f}</td>
      <td>Tk {float(total_will_pay):.0f}</td>
      <td>Tk {float(total_will_receive):.0f}</td>
    </tr></tfoot>
  </table>
  <div class="note">Meal Rate = Total Expense ÷ Total Effective Meals. Effective Meals reflect Meal Count Settings and any Special Meal Entry multiplier in effect for each date.</div>

  <div class="sigs">
    <div class="sig">Manager's Signature</div>
    <div class="sig">Member's Signature</div>
    <div class="sig">Member's Signature</div>
  </div>
  <div class="foot">
    <span>{mess_name} · Confidential — for internal mess accounting use only</span>
    <span>Preview — click Download for the printable PDF</span>
  </div>
{auto_print_script}
</body>
</html>"""
    return HttpResponse(html, content_type='text/html; charset=utf-8')


@xframe_options_sameorigin
@login_required
def export_pdf_en(request):
    member = request.user.member
    if not member.has_perm_code('export_data'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    month = int(request.GET.get('month', get_active_my(request)[0]))
    year  = int(request.GET.get('year',  get_active_my(request)[1]))
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                     Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

    stats = compute_month_stats(month, year)
    buf   = io.BytesIO()
    mess_name = member.mess.name

    PAGE_W, PAGE_H = landscape(A4)
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=8*mm, bottomMargin=16*mm,
                            title=f'{settings.APP_NAME} — Monthly Report {calendar.month_name[month]} {year}',
                            author=settings.APP_NAME)

    # ── Palette ───────────────────────────────────────────────────────────
    NAVY    = colors.HexColor('#161A2E')
    ACCENT  = colors.HexColor('#4F46E5')
    ACCENT2 = colors.HexColor('#7C3AED')
    GREEN   = colors.HexColor('#0F8A3E')
    RED     = colors.HexColor('#D42A2A')
    AMBER   = colors.HexColor('#B45309')
    LIGHT   = colors.HexColor('#F7F8FF')
    SUMM    = colors.HexColor('#F0F1FE')
    GRAY    = colors.HexColor('#8A8AA0')
    LINE    = colors.HexColor('#E1E2F0')

    styles = getSampleStyleSheet()
    brand_style = ParagraphStyle('Brand', fontName='Helvetica-Bold', fontSize=16,
                                  textColor=colors.white, alignment=TA_LEFT, leading=19)
    brand_sub   = ParagraphStyle('BrandSub', fontName='Helvetica', fontSize=8.5,
                                  textColor=colors.HexColor('#B9BAE0'), alignment=TA_LEFT)
    doc_title   = ParagraphStyle('DocTitle', fontName='Helvetica-Bold', fontSize=13,
                                  textColor=colors.white, alignment=TA_RIGHT, leading=16)
    doc_period  = ParagraphStyle('DocPeriod', fontName='Helvetica', fontSize=9.5,
                                  textColor=colors.HexColor('#B9BAE0'), alignment=TA_RIGHT)
    meta_style  = ParagraphStyle('Meta', fontName='Helvetica', fontSize=7.8,
                                  textColor=GRAY, alignment=TA_LEFT)
    meta_style_r = ParagraphStyle('MetaR', fontName='Helvetica', fontSize=7.8,
                                   textColor=GRAY, alignment=TA_RIGHT)
    section_style = ParagraphStyle('Section', fontName='Helvetica-Bold', fontSize=10.5,
                                    textColor=NAVY, alignment=TA_LEFT)

    story = []
    page_w = PAGE_W - 24*mm  # usable width

    # ── Letterhead ──────────────────────────────────────────────────────────
    header_tbl = Table([[
        Paragraph(f'{mess_name}', brand_style),
        Paragraph('MONTHLY FINANCIAL REPORT', doc_title),
    ]], colWidths=[page_w*0.6, page_w*0.4])
    header_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), NAVY),
        ('VALIGN',        (0,0), (-1,-1), 'MIDDLE'),
        ('TOPPADDING',    (0,0), (-1,-1), 10),
        ('BOTTOMPADDING', (0,0), (-1,-1), 2),
        ('LEFTPADDING',   (0,0), (0,0),   12),
        ('RIGHTPADDING',  (1,0), (1,0),   12),
    ]))
    header_sub_tbl = Table([[
        Paragraph(f'{settings.APP_NAME} · Mess ID: {member.mess.code}', brand_sub),
        Paragraph(f'{calendar.month_name[month]} {year}', doc_period),
    ]], colWidths=[page_w*0.6, page_w*0.4])
    header_sub_tbl.setStyle(TableStyle([
        ('BACKGROUND',    (0,0), (-1,-1), NAVY),
        ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ('TOPPADDING',    (0,0), (-1,-1), 0),
        ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ('LEFTPADDING',   (0,0), (0,0),   12),
        ('RIGHTPADDING',  (1,0), (1,0),   12),
    ]))
    story.append(header_tbl)
    story.append(header_sub_tbl)
    # accent underline strip beneath the letterhead
    accent_strip = Table([['']], colWidths=[page_w], rowHeights=[2.4])
    accent_strip.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,-1), ACCENT)]))
    story.append(accent_strip)
    story.append(Spacer(1, 3*mm))

    # ── Meta line (generated by / on, member count) ─────────────────────────
    now = datetime.datetime.now()
    meta_tbl = Table([[
        Paragraph(f"Prepared by: <b>{member.name}</b> ({member.get_role_display()})", meta_style),
        Paragraph(f"Members: <b>{len(stats['per_member'])}</b>", meta_style),
        Paragraph(f"Generated: {now.strftime('%d %b %Y, %I:%M %p')}", meta_style_r),
    ]], colWidths=[page_w*0.42, page_w*0.2, page_w*0.38])
    meta_tbl.setStyle(TableStyle([
        ('TOPPADDING', (0,0), (-1,-1), 0), ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ('LEFTPADDING', (0,0), (0,0), 0), ('RIGHTPADDING', (-1,0), (-1,0), 0),
    ]))
    story.append(meta_tbl)
    story.append(Spacer(1, 3*mm))
    story.append(HRFlowable(width=page_w, thickness=0.6, color=LINE))
    story.append(Spacer(1, 4*mm))

    # ── Summary section ─────────────────────────────────────────────────────
    story.append(Paragraph('FINANCIAL SUMMARY', section_style))
    story.append(Spacer(1, 2.5*mm))

    summ_label = ParagraphStyle('SummL', fontName='Helvetica', fontSize=7.6,
                                 textColor=GRAY, alignment=TA_CENTER)
    summ_value = ParagraphStyle('SummV', fontName='Helvetica-Bold', fontSize=13,
                                 textColor=NAVY, alignment=TA_CENTER, spaceBefore=2)
    fund_bal = float(stats['fund_balance'])
    fund_color = GREEN if fund_bal >= 0 else RED

    cards = [
        ('Meal Rate',           f"Tk {stats['meal_rate']}",              ACCENT),
        ('Total Expense',       f"Tk {stats['total_exp']:.0f}",          RED),
        ('Total Deposit',       f"Tk {stats['total_dep']:.0f}",          GREEN),
        ('Effective Meals',     f"{stats['total_eff']:.1f}",           ACCENT2),
        ('Cook Cost / Member',  f"Tk {stats['cook_cost']}",              AMBER),
        ('Fund Balance',        f"Tk {fund_bal:.0f}",                    fund_color),
    ]
    summ_data = [
        [Paragraph(c[0].upper(), summ_label) for c in cards],
        [Paragraph(c[1], ParagraphStyle('V', parent=summ_value, textColor=c[2])) for c in cards],
    ]
    cw = page_w / 6
    summ_tbl = Table(summ_data, colWidths=[cw]*6, rowHeights=[13, 20])
    summ_style_cmds = [
        ('BACKGROUND',    (0,0), (-1,-1), SUMM),
        ('BOX',           (0,0), (-1,-1), 0.6, LINE),
        ('INNERGRID',     (0,0), (-1,-1), 0.6, LINE),
        ('TOPPADDING',    (0,0), (-1,-1), 5),
        ('BOTTOMPADDING', (0,1), (-1,1), 7),
        ('BOTTOMPADDING', (0,0), (-1,0), 0),
    ]
    for i, c in enumerate(cards):
        summ_style_cmds.append(('LINEABOVE', (i,0), (i,0), 2.2, c[2]))
    summ_tbl.setStyle(TableStyle(summ_style_cmds))
    story.append(summ_tbl)
    story.append(Spacer(1, 5*mm))

    # ── Member breakdown section ────────────────────────────────────────────
    story.append(Paragraph('MEMBER-WISE BREAKDOWN', section_style))
    story.append(Spacer(1, 2.5*mm))

    hdr_style = ParagraphStyle('Hdr', fontName='Helvetica-Bold', fontSize=8.6,
                                textColor=colors.white, alignment=TA_CENTER, leading=10.5)
    cell_c = ParagraphStyle('CC', fontName='Helvetica', fontSize=8.8, alignment=TA_CENTER)
    cell_l = ParagraphStyle('CL', fontName='Helvetica', fontSize=8.8, alignment=TA_LEFT)
    cell_b_g = ParagraphStyle('CG', fontName='Helvetica-Bold', fontSize=8.8,
                               textColor=GREEN, alignment=TA_CENTER)
    cell_b_r = ParagraphStyle('CR', fontName='Helvetica-Bold', fontSize=8.8,
                               textColor=RED, alignment=TA_CENTER)

    headers = ['#', 'Member Name', 'Room', 'Total\nMeal',
               'Meal Cost\n(Tk)', 'Cook Cost\n(Tk)', 'Total Cost\n(Tk)', 'Deposit\n(Tk)',
               'Will Pay\n(Tk)', 'Will Receive\n(Tk)']
    tbl_data = [[Paragraph(h, hdr_style) for h in headers]]

    total_will_pay = Decimal('0')
    total_will_receive = Decimal('0')
    for i, s in enumerate(stats['per_member'], 1):
        bal = float(s['balance'])
        if bal < 0:
            will_pay_p     = Paragraph(f"Tk {abs(bal):.2f}", cell_b_r)
            will_receive_p = Paragraph('—', cell_c)
            total_will_pay += abs(s['balance'])
        else:
            will_pay_p     = Paragraph('—', cell_c)
            will_receive_p = Paragraph(f"Tk {bal:.2f}", cell_b_g)
            total_will_receive += s['balance']
        tbl_data.append([
            Paragraph(str(i), cell_c),
            Paragraph(s['member'].name, cell_l),
            Paragraph(s['member'].room_number or '—', cell_c),
            Paragraph(f"{float(s['eff_meals']):.1f}", cell_c),
            Paragraph(f"Tk {round(float(s['cost'])-float(s['cook_cost']),2):.2f}", cell_c),
            Paragraph(f"Tk {float(s['cook_cost']):.0f}", cell_c),
            Paragraph(f"Tk {float(s['cost']):.2f}", cell_c),
            Paragraph(f"Tk {float(s['deposit']):.0f}", cell_c),
            will_pay_p,
            will_receive_p,
        ])

    # totals row
    tbl_data.append([
        Paragraph('TOTAL', hdr_style), Paragraph('', cell_c), Paragraph('', cell_c),
        Paragraph(f"{float(stats['total_eff']):.1f}", hdr_style),
        Paragraph('', cell_c), Paragraph('', cell_c),
        Paragraph(f"Tk {float(stats['total_exp']):.0f}", hdr_style),
        Paragraph(f"Tk {float(stats['total_dep']):.0f}", hdr_style),
        Paragraph(f"Tk {float(total_will_pay):.0f}", hdr_style),
        Paragraph(f"Tk {float(total_will_receive):.0f}", hdr_style),
    ])

    col_w = [14*mm, 44*mm, 15*mm, 20*mm, 25*mm, 22*mm, 25*mm, 23*mm, 26*mm, 28*mm]
    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    n   = len(stats['per_member'])

    ts = TableStyle([
        ('BACKGROUND',    (0,0),  (-1,0),    ACCENT),
        ('BACKGROUND',    (0,-1), (-1,-1),   NAVY),
        ('ROWBACKGROUNDS',(0,1),  (-1,n),    [colors.white, LIGHT]),
        ('LINEBELOW',     (0,0),  (-1,0),    1, ACCENT2),
        ('GRID',          (0,0),  (-1,-1),   0.35, LINE),
        ('TOPPADDING',    (0,0),  (-1,-1),   5),
        ('BOTTOMPADDING', (0,0),  (-1,-1),   5),
        ('LEFTPADDING',   (0,0),  (-1,-1),   3),
        ('RIGHTPADDING',  (0,0),  (-1,-1),   3),
        ('VALIGN',        (0,0),  (-1,-1),   'MIDDLE'),
    ])
    tbl.setStyle(ts)
    story.append(tbl)
    story.append(Spacer(1, 5*mm))
    story.append(HRFlowable(width=page_w, thickness=0.6, color=LINE))
    story.append(Spacer(1, 2*mm))
    story.append(Paragraph(
        "Meal Rate = Total Expense ÷ Total Effective Meals. "
        "Effective Meals reflect Meal Count Settings and any Special Meal Entry multiplier in effect for each date.",
        ParagraphStyle('FormulaNote', fontName='Helvetica-Oblique', fontSize=7.6,
                        textColor=GRAY, alignment=TA_LEFT)))
    story.append(Spacer(1, 10*mm))

    # ── Signature lines (matches the Bangla PDF's signature block) ──────────
    sig_label = ParagraphStyle('SigLabel', fontName='Helvetica', fontSize=8,
                                textColor=NAVY, alignment=TA_CENTER)
    sig_tbl = Table([[
        Paragraph("Manager's Signature", sig_label),
        Paragraph("Member's Signature", sig_label),
        Paragraph("Member's Signature", sig_label),
    ]], colWidths=[page_w/3]*3, rowHeights=[10*mm])
    sig_tbl.setStyle(TableStyle([
        ('LINEABOVE',  (0,0), (0,0), 1, NAVY),
        ('LINEABOVE',  (1,0), (1,0), 1, NAVY),
        ('LINEABOVE',  (2,0), (2,0), 1, NAVY),
        ('TOPPADDING', (0,0), (-1,-1), 3),
        ('LEFTPADDING',(0,0), (-1,-1), 20),
        ('RIGHTPADDING',(0,0), (-1,-1), 20),
    ]))
    story.append(sig_tbl)

    # ── Footer (page number + confidentiality note on every page) ──────────
    def _footer(canvas, doc_):
        canvas.saveState()
        canvas.setStrokeColor(LINE)
        canvas.setLineWidth(0.5)
        canvas.line(12*mm, 11*mm, PAGE_W-12*mm, 11*mm)
        canvas.setFont('Helvetica', 7)
        canvas.setFillColor(GRAY)
        canvas.drawString(12*mm, 6*mm, f"{mess_name} · Confidential — for internal mess accounting use only")
        canvas.drawRightString(PAGE_W-12*mm, 6*mm, f"Page {doc_.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_footer, onLaterPages=_footer)
    buf.seek(0)
    fname = f"meal_report_en_{month}_{year}.pdf"
    log_action(request.user.member,
               f'Exported PDF (English): {calendar.month_name[month]} {year}',
               request=request)
    disposition = 'inline' if request.GET.get('preview') == '1' else 'attachment'
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'{disposition}; filename="{fname}"'
    return response


# ---------------------------------------------------------------------------
# Export: PDF (Bangla)  — all labels bold, now shows mess name
# ---------------------------------------------------------------------------

@login_required
def export_pdf_bn(request):
    member = request.user.member
    if not member.has_perm_code('export_data'):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    month = int(request.GET.get('month', get_active_my(request)[0]))
    year  = int(request.GET.get('year',  get_active_my(request)[1]))
    is_preview = request.GET.get('preview') == '1'
    # No headless-browser PDF renderer is used for Bangla (ReportLab can't
    # shape Bangla script correctly, and a server-side browser dependency
    # caused persistent setup trouble) — so both the default (Download)
    # and the explicit Print mode open this page and trigger the
    # browser's own print dialog, letting the user choose "Save as PDF".
    # Only the preview (shown inside the modal's iframe) skips this.
    script_block = """<script>
window.onload = function() {
  setTimeout(function() { window.print(); }, 700);
};
</script>""" if not is_preview else ''

    # Build self-contained HTML → return as printable HTML page
    # (ReportLab has no Bangla-font support out of the box; the browser's
    #  print-to-PDF path handles Unicode perfectly via Hind Siliguri.)
    stats = compute_month_stats(month, year)
    mess_name = member.mess.name
    today = datetime.date.today()
    print_date = f"{_bn_num(today.day)} {BN_MONTHS[today.month-1]} {_bn_num(today.year)}"
    month_label = _bn_month_year(month, year)

    # build rows
    rows_html = ''
    for i, s in enumerate(stats['per_member'], 1):
        bal       = float(s['balance'])
        pabe      = _bn_num(f"{abs(bal):.2f}") if bal >= 0 else '—'
        dibe      = _bn_num(f"{abs(bal):.2f}") if bal < 0  else '—'
        bg        = '#f0f0f0' if i % 2 == 0 else '#fff'
        name      = _member_display_name(s['member'], 'bn')
        eff_meals = _bn_num(f"{float(s['eff_meals']):.1f}")
        meal_cost = _bn_num(f"{round(float(s['cost'])-float(s['cook_cost']),2):.2f}")
        cook_cost = _bn_num(f"{float(s['cook_cost']):.0f}")
        total_c   = _bn_num(f"{float(s['cost']):.2f}")
        deposit   = _bn_num(f"{float(s['deposit']):.0f}")
        rows_html += (
            f'<tr style="background:{bg}">'
            f'<td class="sl">{_bn_num(i)}</td>'
            f'<td class="name">{name}</td>'
            f'<td>{eff_meals}</td>'
            f'<td>{meal_cost}</td>'
            f'<td>{cook_cost}</td>'
            f'<td><b>{total_c}</b></td>'
            f'<td>{deposit}</td>'
            f'<td class="pabe">{pabe}</td>'
            f'<td class="dibe">{dibe}</td>'
            f'</tr>'
        )

    fund_bal   = float(stats['fund_balance'])
    fund_color = '#1a7a1a' if fund_bal >= 0 else '#cc0000'

    html = f"""<!DOCTYPE html>
<html lang="bn">
<head>
<meta charset="UTF-8">
<title>মাসিক হিসাব — {month_label}</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=Hind+Siliguri:wght@400;600;700&display=swap" rel="stylesheet">
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
@page{{size:A4 landscape;margin:8mm 8mm}}
body{{
  font-family:'Hind Siliguri','Noto Sans Bengali',Arial,sans-serif;
  font-weight:700;
  font-size:10pt;color:#000;background:#fff;padding:4mm 5mm
}}
/* ALL labels bold — universal rule */
*{{font-weight:700}}
.hdr{{text-align:center;background:#1e2235;color:#fff;padding:6px 10px 5px;border-radius:3px;margin-bottom:6px}}
.hdr-t1{{font-size:14pt;margin-bottom:2px}}
.hdr-t2{{font-size:10pt;color:#aaaacc}}
.sum{{display:flex;gap:4px;margin-bottom:6px}}
.sc{{flex:1;border:1.5px solid #aaa;border-radius:2px;padding:4px 5px;text-align:center;background:#eef0ff}}
.sc-l{{font-size:8pt;color:#333366;margin-bottom:1px}}
.sc-v{{font-size:11pt}}
table{{width:100%;border-collapse:collapse;font-size:9.5pt}}
th{{background:#4f46e5;color:#fff;padding:5px 4px;text-align:center;
    border:1px solid #000;line-height:1.3}}
th.name{{text-align:left;padding-left:7px}}
td{{padding:4px 4px;border:1px solid #bbb;text-align:center;vertical-align:middle;line-height:1.3}}
td.name{{text-align:left;padding-left:7px}}
td.sl{{font-size:8.5pt;color:#333;width:22px}}
tr:nth-child(even){{background:#f0f0f0}}
tr:nth-child(odd){{background:#fff}}
tfoot tr td{{background:#1e2235;color:#fff;border:1px solid #000;padding:5px 4px;text-align:center}}
tfoot tr td.name{{text-align:left;padding-left:7px}}
.pabe{{color:#1a7a1a}}
.dibe{{color:#cc0000}}
.foot{{margin-top:6px;border-top:1px solid #aaa;padding-top:4px;
       display:flex;justify-content:space-between;font-size:8pt;color:#444}}
.sigs{{display:flex;justify-content:space-between;margin-top:14px;padding:0 8px}}
.sig{{text-align:center;padding-top:4px;border-top:1.5px solid #000;
      font-size:8pt;min-width:90px}}
</style>
</head>
<body>
<div class="hdr">
  <div class="hdr-t1">মাসিক হিসাব রিপোর্ট</div>
  <div class="hdr-t2">{mess_name} · {month_label}</div>
</div>
<div class="sum">
  <div class="sc"><div class="sc-l">মোট বাজার খরচ</div><div class="sc-v">৳{_bn_num(f"{float(stats['total_exp']):.0f}")}</div></div>
  <div class="sc"><div class="sc-l">মোট জমা</div><div class="sc-v">৳{_bn_num(f"{float(stats['total_dep']):.0f}")}</div></div>
  <div class="sc" style="border-color:{fund_color}"><div class="sc-l">ফান্ড ব্যালেন্স</div><div class="sc-v" style="color:{fund_color}">৳{_bn_num(f"{fund_bal:.0f}")}</div></div>
  <div class="sc"><div class="sc-l">মিল রেট</div><div class="sc-v">৳{_bn_num(stats['meal_rate'])}</div></div>
  <div class="sc"><div class="sc-l">রান্না খরচ/জন</div><div class="sc-v">৳{_bn_num(stats['cook_cost'])}</div></div>
  <div class="sc"><div class="sc-l">মোট কার্যকর মিল</div><div class="sc-v">{_bn_num(f"{float(stats['total_eff']):.1f}")}</div></div>
</div>
<table>
  <thead>
    <tr>
      <th style="width:22px">ক্র:</th>
      <th class="name">নাম</th>
      <th>মোট মিল</th>
      <th>মিল খরচ (৳)</th>
      <th>রান্না খরচ (৳)</th>
      <th>মোট খরচ (৳)</th>
      <th>জমা (৳)</th>
      <th>পাবে ✓</th>
      <th>দিবে ✗</th>
    </tr>
  </thead>
  <tbody>{rows_html}</tbody>
  <tfoot>
    <tr>
      <td colspan="2" class="name">মোট</td>
      <td>{_bn_num(f"{float(stats['total_eff']):.1f}")}</td>
      <td colspan="2"></td>
      <td>৳{_bn_num(f"{float(stats['total_exp']):.0f}")}</td>
      <td>৳{_bn_num(f"{float(stats['total_dep']):.0f}")}</td>
      <td colspan="2"></td>
    </tr>
  </tfoot>
</table>
<div class="foot">
  <span>মুদ্রণের তারিখ: {print_date}</span>
  <span>মিল রেট = মোট বাজার ÷ মোট কার্যকর মিল</span>
</div>
<div class="sigs">
  <div class="sig">ম্যানেজারের স্বাক্ষর</div>
  <div class="sig">সদস্যের স্বাক্ষর</div>
  <div class="sig">সদস্যের স্বাক্ষর</div>
</div>
{script_block}
</body>
</html>"""

    log_action(request.user.member,
               f'Exported PDF (Bangla): {calendar.month_name[month]} {year}',
               request=request)

    return HttpResponse(html, content_type='text/html; charset=utf-8')


# ---------------------------------------------------------------------------
# Analytics
# ---------------------------------------------------------------------------

@login_required
def analytics(request):
    today  = datetime.date.today()
    # Preloaded once for the whole 12-month trend loop below, instead of
    # a fresh DayConfig + MealCountSettings query per MealMark row.
    day_config_map, weights = MealMark.preload_calc_context(request.user.member.mess)
    trend  = []
    for i in range(11, -1, -1):
        mn = ((today.month - 1 - i) % 12) + 1
        yr = today.year if i <= today.month - 1 else today.year - 1
        eff = sum((mk.effective_count(day_config_map, weights)
                  for mk in MealMark.objects.filter(date__month=mn, date__year=yr).select_related('member__mess')),
                  Decimal('0'))
        exp = Expense.objects.filter(date__month=mn, date__year=yr
              ).aggregate(t=Sum('amount'))['t'] or 0
        dep = Deposit.objects.filter(date__month=mn, date__year=yr
              ).aggregate(t=Sum('amount'))['t'] or 0
        trend.append({
            'label': f"{calendar.month_abbr[mn]}'{str(yr)[-2:]}",
            'eff': float(eff), 'exp': float(exp), 'dep': float(dep),
        })
    cat_data  = (Expense.objects.filter(date__year=today.year)
                 .values('category').annotate(t=Sum('amount')).order_by('-t'))
    cat_labels = [c['category'] for c in cat_data]
    cat_vals   = [float(c['t']) for c in cat_data]
    top_dep    = (Deposit.objects.filter(date__year=today.year)
                  .values('member__name').annotate(t=Sum('amount')).order_by('-t')[:5])
    return render(request, 'reports/analytics.html', {
        'trend_json':    json.dumps(trend),
        'cat_labels':    json.dumps(cat_labels),
        'cat_vals':      json.dumps(cat_vals),
        'top_depositors': list(top_dep),
    })
